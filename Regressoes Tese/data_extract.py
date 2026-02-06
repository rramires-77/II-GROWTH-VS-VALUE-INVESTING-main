# NÃO RODAR AQUI. COPIE PARA O SEU AMBIENTE.
# Lê percentuais do RAW_TEXT e escreve, em ordem, na 2ª coluna (B) da planilha.

from openpyxl import load_workbook
import re
import os

# --------- Caminhos ---------
ARQ = "/Users/zzxd/Projetos/progresso/Rogerio/artigo-2/dados/portugal_psi20_data_template.xlsx"
SAIDA = "/Users/zzxd/Projetos/progresso/Rogerio/artigo-2/dados/portugal_psi20_data_filled.xlsx"
NOME_ABA = None  # deixe None para usar a ativa; ou coloque o nome da aba como string

# --------- Cole aqui o texto bruto (na ordem que deseja gravar) ---------
RAW_TEXT = """
ALTRISGPS 11% 8% 10% D% 10% 2% 75%
B.COM.PORTUGUES 2% 0% 0% 2% 1% 1% 3%
CORTICEIRA AMORIM 3% 2% 3% 2% 2% 3% 4%
CTT CORREIOS PORT 7% 14% 13% 3% 0% 2% 4%
EDP 8% 8% 8% 6% 4% 5% 5%
EDP RENOVAVEIS 2% 2% 2% 1% 1% 1% 1%
GALP ENERGIA-NOM 3% 4% 4% 5% 6% 10% 1%
IBERSOL. SGPS 3% 1% 1% 1% 0% 0% 2%
JMARTINS, SGPS 2% 4% 6% 2% 3% 2% 4%
MOTA ENGIL 13% 7% 13% 9% 3% 9% 9%
NOS, SGPS 3% 4% 5% 7% 10% 8% 4%
NOVABASE, SGPS 6% D% 8% 20% 0% 0% 0%
PHAROL 14% 0% 0% 0% 0% 0% 0%
RAMADA 4% 2% 29% 10% 0% 8% 9%
REN 7% 5% 7% 6% 7% 1% 9%
SEMAPA 9% 9% 10% 2% 6% 9% M%
SONAE 0% 4% 6% 17% 13% 5% 9%
THE NAVIGATOR CO. 9% 8% 7% 12% 0% 6% 5%
""".strip()

# --------- Extração simples dos percentuais, na ordem ---------
# Captura tokens tipo "12%", "7,5%" etc.
tokens = re.findall(r"-?\d+(?:[.,]\d+)?%", RAW_TEXT)

# --------- Converte para número (sem %) ---------
def pct_to_float(tok: str) -> float:
    t = tok.strip().replace("%", "").replace(",", ".")
    return float(t)

values = [pct_to_float(t) for t in tokens]

# --------- Abre a planilha ---------
wb = load_workbook(ARQ)
ws = wb[NOME_ABA] if NOME_ABA and NOME_ABA in wb.sheetnames else wb.active

# --------- Escreve na coluna C (3ª coluna), a partir da linha 2 ---------
row = 2
col = 3  # C
for i, v in enumerate(values, start=1):
    ws.cell(row=row, column=col).value = v
    row += 1  # avança uma linha
    if i % 7 == 0:  # fim do bloco de 7 → pula +2 linhas
        row += 2


# --------- Salva em novo arquivo ---------
wb.save(SAIDA)
print(f"Gravado em: {SAIDA}\nTotal de valores escritos: {len(values)}")
